from fastapi import FastAPI, Request, UploadFile, File
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook

import database


app = FastAPI()

templates = Jinja2Templates(directory="templates")


# Initialisation base
database.init_db()


# Page accueil
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="index.html"
    )


# Page scan
@app.get("/scan-page", response_class=HTMLResponse)
def scan_page(request: Request):
    return templates.TemplateResponse(
        request=request,
        name="scan.html"
    )


# Scanner un ticket
@app.get("/scan/{ticket_id}")
def scan(ticket_id: str):

    ticket = database.get_ticket(ticket_id)

    if ticket is None:
        return {
            "status": "inconnu"
        }


    # déjà utilisé
    if ticket[2]:
        return {
            "status": "deja_utilise",
            "nom": ticket[1]
        }


    # valide le ticket
    database.validate_ticket(ticket_id)


    return {
        "status": "valide",
        "nom": ticket[1]
    }



# Page import Excel
@app.get("/import", response_class=HTMLResponse)
async def import_page(request: Request):

    return templates.TemplateResponse(
        request=request,
        name="import.html"
    )



# Import fichier Excel
@app.post("/import")
async def import_excel(file: UploadFile = File(...)):

    workbook = load_workbook(file.file)

    sheet = workbook.active


    compteur = 0


    for row in sheet.iter_rows(values_only=True):

        if row[0] is None:
            continue


        numero = str(row[0])

        nom = ""

        if len(row) > 1 and row[1]:
            nom = str(row[1])


        database.add_ticket(numero, nom)

        compteur += 1


    return {
        "message": "Tickets importés",
        "nombre": compteur
    }